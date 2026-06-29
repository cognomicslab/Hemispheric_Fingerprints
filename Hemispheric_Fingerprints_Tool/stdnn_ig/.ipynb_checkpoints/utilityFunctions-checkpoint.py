import numpy as np
import math
import os
from captum.attr import IntegratedGradients
from captum.attr import DeepLift
from nilearn import plotting,image,surface,datasets
from scipy.spatial.distance import squareform, pdist
import random
import sys


def reshapeData(data):
    # print(data.shape)
    no_subjs, no_ts, no_channels = data.shape
    # Reshape data to no_subjs, no_channels, no_ts
    data_reshape = np.empty((no_subjs, no_channels, no_ts))
    for subj in np.arange(no_subjs):
        x_subj = data[subj, :, :]
        x_subj = np.transpose(x_subj)
        data_reshape[subj, :, :] = x_subj
    # print(data_reshape.shape)
    return data_reshape

def prepare_data_sliding_window(data, labels,window_size, step):
    ''' Function to create windowed data'''
    Nsubjs,N,Nchannels = data.shape
    width = int(np.floor(window_size / 2.0))
    labels_window = []
    window_data_list=[]
    for subj in np.arange(Nsubjs):
        #print("subject = ",subj)
        for k in range(width, N - width - 1, step):
            x = data[subj,k - width: k + width,:]
            x = np.expand_dims(x,axis=0)
            window_data_list.append(x)
            labels_window.append(labels[subj])
    window_data = np.vstack(window_data_list)
    return (window_data,labels_window)


def write_excel_file(accuracy, precision, recall, f1, excel_file, model_ss, test_ss,auc_values):
    print("Test。。。。。。。")
    print('accuracy:',accuracy)
    print('precision:',precision)
    print('recall:',recall)
    print('f1:',f1)
    print('auc_values:',auc_values)
    if not os.path.exists(excel_file): # excel file name
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "trained_%s_tested_%s" % (model_ss, test_ss) # excel sheet name
    else:
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        ws1 = wb.create_sheet(title="trained_%s_tested_%s" % (model_ss, test_ss))

    ws1.append(["Session", "Fold Number", "Accuracy", "Precision", "Recall", "F1-score", "AUC"])

    def format_no_round(value):
        value = int(value*1000)/1000
        return f"{value:.3f}"

    for idx in range(len(accuracy)):
        if idx == 0:
            ws1.append([test_ss, "%01d" % (idx+1), 
                       format_no_round(accuracy[idx]),
                       format_no_round(precision[idx]),
                       format_no_round(recall[idx]),
                       format_no_round(f1[idx]),
                       format_no_round(auc_values[idx])])
        else:
            ws1.append(["", "%01d" % (idx + 1),
                       format_no_round(accuracy[idx]),
                       format_no_round(precision[idx]),
                       format_no_round(recall[idx]),
                       format_no_round(f1[idx]),
                       format_no_round(auc_values[idx])])
    
    # ws1.append(["", "Avg (Std)",
    #            format_no_round(np.mean(accuracy)) + " (" + format_no_round(np.std(accuracy)) + ")",
    #            format_no_round(np.mean(precision)) + " (" + format_no_round(np.std(precision)) + ")",
    #            format_no_round(np.mean(recall)) + " (" + format_no_round(np.std(recall)) + ")",
    #            format_no_round(np.mean(f1)) + " (" + format_no_round(np.std(f1)) + ")",
    #            format_no_round(np.mean(auc_values)) + " (" + format_no_round(np.std(auc_values)) + ")"])
    print(format_no_round(np.mean(accuracy)))
    print(format_no_round(np.mean(precision)))
    print(format_no_round(np.mean(recall)))
    print(format_no_round(np.mean(f1)))
    print(format_no_round(np.mean(auc_values)))
    print(format_no_round(0.99999902))
    
    wb.save(filename=excel_file)
    
# target should be 0 for left and 1 for right 
def getInputAttributions(model, input_tensor,target):
    ig = IntegratedGradients(model)
    input_tensor.requires_grad_()
    attr, delta = ig.attribute(input_tensor, target=target, return_convergence_delta=True)
    attr = attr.cpu().detach().numpy()
   # attr = attr.detach().numpy()
    return attr


def determine_features(data_file,group_label,percentile):

    data = np.load(data_file)
    # 筛选特定组别的数据
    group_features = data['features'][np.where(data['labels'] == group_label)]
    # 计算中位数:对筛选后的特征数据 group_features 在时间点维度（第二维）上计算中位数。结果是一个 N×192的数组，其中每个元素是特定被试在特定脑区的中位数。
    medians = np.median(group_features, axis=2) 
    # 计算平均绝对中位数 计算这些绝对值在所有被试（subjects）上的均值。结果是一个长度为192的数组，其中每个元素是特定脑区的平均绝对中位数。
    mean_across_subj = np.mean(np.abs(medians), axis=0) 
    # 使用 np.percentile 计算 mean_across_subj 在所有脑区上指定 percentile 的阈值。使用 np.where 找出所有大于或等于这个阈值的脑区索引。
    percentiles = np.where(np.abs(mean_across_subj) >= np.percentile(np.abs(mean_across_subj),percentile))
    features_idcs = percentiles[0] # 包含所有大于或等于阈值的脑区索引的数组。
    features = mean_across_subj # 包含每个脑区的平均特征分数（即平均绝对中位数）的数组

    return features_idcs,features


